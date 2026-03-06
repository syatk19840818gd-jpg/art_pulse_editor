#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import load_workbook


def choose_xlsx_path() -> Path:
    p1 = Path("/mnt/data/rag_gellery_breakdown_master.xlsx")
    if p1.exists():
        return p1
    p2 = Path("data/gallery_lists/rag_gellery_breakdown_master.xlsx")
    if p2.exists():
        return p2
    raise FileNotFoundError("rag_gellery_breakdown_master.xlsx not found")


def normalize_gallery_key(raw: str) -> str:
    s = unicodedata.normalize("NFKC", str(raw or ""))
    s = s.lower().strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.split())


def canonical_url(url: str) -> str:
    if not url:
        return ""
    parts = urlsplit(url.strip())
    scheme = "https"
    netloc = (parts.netloc or "").lower()
    if netloc.startswith("www."):
        netloc = netloc[4:]
    path = parts.path or "/"
    while "//" in path:
        path = path.replace("//", "/")
    if len(path) > 1 and path.endswith("/"):
        path = path[:-1]
    tracking = {"fbclid", "gclid", "mc_cid", "mc_eid", "utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content"}
    q = [(k, v) for k, v in parse_qsl(parts.query, keep_blank_values=True) if k.lower() not in tracking]
    query = urlencode(sorted(q), doseq=True)
    return urlunsplit((scheme, netloc, path, query, ""))


def read_jsonl(path: Path):
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except Exception:
                continue


def text_non_empty(v: object) -> bool:
    return bool(str(v or "").strip())


class GalleryStats:
    __slots__ = (
        "display",
        "artist_img_keys",
        "artist_txt_keys",
        "artist_img_count",
        "artist_txt_count",
        "exh_img_keys",
        "exh_txt_keys",
        "exh_img_count",
        "exh_txt_count",
    )

    def __init__(self):
        self.display = ""
        self.artist_img_keys = set()
        self.artist_txt_keys = set()
        self.artist_img_count = 0
        self.artist_txt_count = 0
        self.exh_img_keys = set()
        self.exh_txt_keys = set()
        self.exh_img_count = 0
        self.exh_txt_count = 0


def get_gallery_name(row: dict) -> str:
    return str(row.get("gallery_name_en") or row.get("gallery_name") or row.get("gallery") or "")


def build_agg() -> dict[str, dict[str, dict[str, float | int | None | str]]]:
    base = Path("data/phase1_seed10")
    fair_files = {
        "frieze-london": {
            "artist_text": base / "raw/artists_frieze_london_2025.jsonl",
            "artist_img": base / "derived/artist_works_images_frieze_london.jsonl",
            "exh_text": base / "raw/exhibitions_frieze_london_2025.jsonl",
            "exh_img": base / "derived/exhibitions_images_frieze_london_2025.jsonl",
        },
        "liste": {
            "artist_text": base / "raw/artists_liste_2025.jsonl",
            "artist_img": base / "derived/artist_works_images_liste.jsonl",
            "exh_text": base / "raw/exhibitions_liste_2025.jsonl",
            "exh_img": base / "derived/exhibitions_images_liste_2025.jsonl",
        },
    }
    stats = {fair: defaultdict(GalleryStats) for fair in fair_files}

    for fair, files in fair_files.items():
        d = stats[fair]
        for row in read_jsonl(files["artist_text"]):
            g = get_gallery_name(row)
            if not g:
                continue
            gk = normalize_gallery_key(g)
            rec = d[gk]
            if not rec.display:
                rec.display = g
            if text_non_empty(row.get("text")):
                key = canonical_url(str(row.get("source_url") or ""))
                if key:
                    rec.artist_txt_keys.add(key)
                    rec.artist_txt_count += 1

        for row in read_jsonl(files["artist_img"]):
            g = get_gallery_name(row)
            if not g:
                continue
            gk = normalize_gallery_key(g)
            rec = d[gk]
            if not rec.display:
                rec.display = g
            paths = [x for x in (row.get("works_image_local_paths") or []) if str(x).strip()]
            urls = [x for x in (row.get("works_image_urls") or []) if str(x).strip()]
            img_cnt = len(paths) if paths else len(urls)
            if img_cnt <= 0:
                continue
            key = canonical_url(str(row.get("source_url") or ""))
            if key:
                rec.artist_img_keys.add(key)
            rec.artist_img_count += img_cnt

        for row in read_jsonl(files["exh_text"]):
            g = get_gallery_name(row)
            if not g:
                continue
            gk = normalize_gallery_key(g)
            rec = d[gk]
            if not rec.display:
                rec.display = g
            if text_non_empty(row.get("text")):
                key = canonical_url(str(row.get("source_url") or ""))
                if key:
                    rec.exh_txt_keys.add(key)
                    rec.exh_txt_count += 1

        if files["exh_img"].exists():
            for row in read_jsonl(files["exh_img"]):
                g = get_gallery_name(row)
                if not g:
                    continue
                gk = normalize_gallery_key(g)
                rec = d[gk]
                if not rec.display:
                    rec.display = g
                key = canonical_url(str(row.get("source_url") or ""))
                if key:
                    rec.exh_img_keys.add(key)
                if str(row.get("local_path") or "").strip() or str(row.get("image_url") or "").strip():
                    rec.exh_img_count += 1

    out: dict[str, dict[str, dict[str, float | int | None | str]]] = {}
    for fair, d in stats.items():
        rows = {}
        for gk, rec in d.items():
            a_union = rec.artist_img_keys | rec.artist_txt_keys
            a_match = rec.artist_img_keys & rec.artist_txt_keys
            e_union = rec.exh_img_keys | rec.exh_txt_keys
            e_match = rec.exh_img_keys & rec.exh_txt_keys
            rows[gk] = {
                "gallery_name": rec.display or gk,
                "artist_union": len(a_union),
                "artist_match": len(a_match),
                "artist_rate": (100.0 * len(a_match) / len(a_union)) if a_union else None,
                "artist_img_unique": len(rec.artist_img_keys),
                "artist_img_count": rec.artist_img_count,
                "artist_txt_unique": len(rec.artist_txt_keys),
                "artist_txt_count": rec.artist_txt_count,
                "exh_union": len(e_union),
                "exh_match": len(e_match),
                "exh_rate": (100.0 * len(e_match) / len(e_union)) if e_union else None,
                "exh_img_count": rec.exh_img_count,
                "exh_txt_count": rec.exh_txt_count,
            }
        out[fair] = rows
    return out


def main() -> None:
    xlsx = choose_xlsx_path()
    wb = load_workbook(xlsx)
    run_id = datetime.now(timezone.utc).strftime("TASK_DIAG_ANCAPOTERASU_02_%Y%m%dT%H%M%SZ")
    today = datetime.now().date()

    cols = {
        "gallery": 1,
        "artist_union": 2,
        "artist_match": 3,
        "artist_rate": 4,
        "artist_img_unique": 5,
        "artist_img_count": 6,
        "artist_txt_unique": 7,
        "artist_txt_count": 8,
        "exh_union": 9,
        "exh_match": 10,
        "exh_rate": 11,
        "exh_img_count": 12,
        "exh_txt_count": 13,
        "updated_at": 15,
        "run_id": 16,
    }

    agg = build_agg()
    anca_before = None
    anca_after = None
    non_ascii_checks: list[dict] = []
    sheet_stats: dict[str, dict[str, int]] = {}

    for sheet in ("frieze-london", "liste"):
        ws = wb[sheet]
        existing = {}
        r = 2
        while True:
            g = ws.cell(r, cols["gallery"]).value
            if g is None or str(g).strip() == "":
                break
            gk = normalize_gallery_key(str(g))
            if gk not in existing:
                existing[gk] = r
            if sheet == "liste" and "anca" in gk:
                anca_before = {
                    "row": r,
                    "gallery_name": str(g),
                    "exh_union": ws.cell(r, cols["exh_union"]).value,
                    "exh_match": ws.cell(r, cols["exh_match"]).value,
                    "exh_rate": ws.cell(r, cols["exh_rate"]).value,
                    "exh_img_count": ws.cell(r, cols["exh_img_count"]).value,
                    "exh_txt_count": ws.cell(r, cols["exh_txt_count"]).value,
                }
            r += 1
        next_row = r

        overwritten = 0
        added = 0
        for gk, vals in sorted(agg[sheet].items(), key=lambda kv: str(kv[1]["gallery_name"])):
            if gk in existing:
                rr = existing[gk]
                overwritten += 1
            else:
                rr = next_row
                ws.cell(rr, cols["gallery"]).value = vals["gallery_name"]
                next_row += 1
                added += 1

            ws.cell(rr, cols["artist_union"]).value = vals["artist_union"]
            ws.cell(rr, cols["artist_match"]).value = vals["artist_match"]
            ws.cell(rr, cols["artist_rate"]).value = vals["artist_rate"]
            ws.cell(rr, cols["artist_img_unique"]).value = vals["artist_img_unique"]
            ws.cell(rr, cols["artist_img_count"]).value = vals["artist_img_count"]
            ws.cell(rr, cols["artist_txt_unique"]).value = vals["artist_txt_unique"]
            ws.cell(rr, cols["artist_txt_count"]).value = vals["artist_txt_count"]
            ws.cell(rr, cols["exh_union"]).value = vals["exh_union"]
            ws.cell(rr, cols["exh_match"]).value = vals["exh_match"]
            ws.cell(rr, cols["exh_rate"]).value = vals["exh_rate"]
            ws.cell(rr, cols["exh_img_count"]).value = vals["exh_img_count"]
            ws.cell(rr, cols["exh_txt_count"]).value = vals["exh_txt_count"]
            ws.cell(rr, cols["updated_at"]).value = today
            ws.cell(rr, cols["run_id"]).value = run_id

        sheet_stats[sheet] = {"overwritten": overwritten, "added": added, "total": next_row - 2}

        # collect at most 2 non-ascii check rows per workbook after update
        if len(non_ascii_checks) < 2:
            for rr in range(2, next_row):
                g = ws.cell(rr, cols["gallery"]).value
                if not g:
                    continue
                gs = str(g)
                if any(ord(ch) > 127 for ch in gs):
                    non_ascii_checks.append(
                        {
                            "sheet": sheet,
                            "gallery_name": gs,
                            "exh_img_count": ws.cell(rr, cols["exh_img_count"]).value,
                            "exh_txt_count": ws.cell(rr, cols["exh_txt_count"]).value,
                            "exh_match": ws.cell(rr, cols["exh_match"]).value,
                        }
                    )
                    if len(non_ascii_checks) >= 2:
                        break

    # after values for Anca
    ws_liste = wb["liste"]
    for rr in range(2, ws_liste.max_row + 1):
        g = ws_liste.cell(rr, cols["gallery"]).value
        if g is None or str(g).strip() == "":
            break
        if "anca" in normalize_gallery_key(str(g)):
            anca_after = {
                "row": rr,
                "gallery_name": str(g),
                "exh_union": ws_liste.cell(rr, cols["exh_union"]).value,
                "exh_match": ws_liste.cell(rr, cols["exh_match"]).value,
                "exh_rate": ws_liste.cell(rr, cols["exh_rate"]).value,
                "exh_img_count": ws_liste.cell(rr, cols["exh_img_count"]).value,
                "exh_txt_count": ws_liste.cell(rr, cols["exh_txt_count"]).value,
            }
            break

    wb.save(xlsx)

    log_path = Path("data/phase1_seed10/logs/rag_gallery_breakdown_update_task_diag_ancapoterasu_02.md")
    log_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# TASK_DIAG_ANCAPOTERASU_02",
        f"- run_id: `{run_id}`",
        f"- xlsx: `{xlsx}`",
        "- 修正要点: gallery照合キーを `NFKC -> lower -> trim/collapse spaces -> NFKD+combining除去` に統一。",
        "- 影響: `ş/ș` などのダイアクリティカル差を照合時に吸収（表示名は変更しない）。",
        "",
        "## Anca before/after (liste)",
        f"- before: {json.dumps(anca_before, ensure_ascii=False)}",
        f"- after:  {json.dumps(anca_after, ensure_ascii=False)}",
        "",
        "## Non-ASCII quick checks (max 2)",
        json.dumps(non_ascii_checks, ensure_ascii=False, indent=2),
        "",
        "- 書式保持: 値セルのみ更新（列/結合/色/罫線/列幅/条件付き書式は非変更）。",
    ]
    log_path.write_text("\n".join(lines), encoding="utf-8")

    print(
        json.dumps(
            {
                "run_id": run_id,
                "sheet_stats": sheet_stats,
                "anca_before": anca_before,
                "anca_after": anca_after,
                "non_ascii_checks": non_ascii_checks,
                "log_path": str(log_path),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()

