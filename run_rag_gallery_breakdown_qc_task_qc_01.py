#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import load_workbook


def normalize_gallery_key(raw: str) -> str:
    s = unicodedata.normalize("NFKC", str(raw or ""))
    s = s.lower().strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.split())


def canonical_url(url: str) -> str:
    if not url:
        return ""
    parts = urlsplit(url.strip())
    scheme = "https"
    netloc = (parts.netloc or "").lower()
    if netloc.startswith("www."):
        netloc = netloc[4:]
    path = parts.path or "/"
    while "//" in path:
        path = path.replace("//", "/")
    if len(path) > 1 and path.endswith("/"):
        path = path[:-1]
    tracking = {
        "fbclid",
        "gclid",
        "mc_cid",
        "mc_eid",
        "utm_source",
        "utm_medium",
        "utm_campaign",
        "utm_term",
        "utm_content",
    }
    query_items = [(k, v) for k, v in parse_qsl(parts.query, keep_blank_values=True) if k.lower() not in tracking]
    return urlunsplit((scheme, netloc, path, urlencode(sorted(query_items), doseq=True), ""))


def read_jsonl(path: Path):
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except Exception:
                continue


def text_non_empty(v: object) -> bool:
    return bool(str(v or "").strip())


def get_gallery(row: dict) -> str:
    return str(row.get("gallery_name_en") or row.get("gallery_name") or row.get("gallery") or "")


class Stat:
    __slots__ = (
        "artist_img_keys",
        "artist_txt_keys",
        "artist_img_count",
        "artist_txt_count",
        "exh_img_keys",
        "exh_txt_keys",
        "exh_img_count",
        "exh_txt_count",
    )

    def __init__(self):
        self.artist_img_keys = set()
        self.artist_txt_keys = set()
        self.artist_img_count = 0
        self.artist_txt_count = 0
        self.exh_img_keys = set()
        self.exh_txt_keys = set()
        self.exh_img_count = 0
        self.exh_txt_count = 0


def build_formal_agg() -> dict[str, dict[str, dict[str, float | int | None]]]:
    base = Path("data/phase1_seed10")
    files = {
        "frieze-london": {
            "artist_text": base / "raw/artists_frieze_london_2025.jsonl",
            "artist_img": base / "derived/artist_works_images_frieze_london.jsonl",
            "exh_text": base / "raw/exhibitions_frieze_london_2025.jsonl",
            "exh_img": base / "derived/exhibitions_images_frieze_london_2025.jsonl",
        },
        "liste": {
            "artist_text": base / "raw/artists_liste_2025.jsonl",
            "artist_img": base / "derived/artist_works_images_liste.jsonl",
            "exh_text": base / "raw/exhibitions_liste_2025.jsonl",
            "exh_img": base / "derived/exhibitions_images_liste_2025.jsonl",
        },
    }
    stats = {fair: defaultdict(Stat) for fair in files}
    for fair, f in files.items():
        d = stats[fair]
        for row in read_jsonl(f["artist_text"]):
            g = get_gallery(row)
            if not g:
                continue
            gk = normalize_gallery_key(g)
            if text_non_empty(row.get("text")):
                key = canonical_url(str(row.get("source_url") or ""))
                if key:
                    d[gk].artist_txt_keys.add(key)
                    d[gk].artist_txt_count += 1
        for row in read_jsonl(f["artist_img"]):
            g = get_gallery(row)
            if not g:
                continue
            gk = normalize_gallery_key(g)
            paths = [x for x in (row.get("works_image_local_paths") or []) if str(x).strip()]
            urls = [x for x in (row.get("works_image_urls") or []) if str(x).strip()]
            image_count = len(paths) if paths else len(urls)
            if image_count <= 0:
                continue
            key = canonical_url(str(row.get("source_url") or ""))
            if key:
                d[gk].artist_img_keys.add(key)
            d[gk].artist_img_count += image_count
        for row in read_jsonl(f["exh_text"]):
            g = get_gallery(row)
            if not g:
                continue
            gk = normalize_gallery_key(g)
            if text_non_empty(row.get("text")):
                key = canonical_url(str(row.get("source_url") or ""))
                if key:
                    d[gk].exh_txt_keys.add(key)
                    d[gk].exh_txt_count += 1
        if f["exh_img"].exists():
            for row in read_jsonl(f["exh_img"]):
                g = get_gallery(row)
                if not g:
                    continue
                gk = normalize_gallery_key(g)
                key = canonical_url(str(row.get("source_url") or ""))
                if key:
                    d[gk].exh_img_keys.add(key)
                if str(row.get("local_path") or "").strip() or str(row.get("image_url") or "").strip():
                    d[gk].exh_img_count += 1
    out: dict[str, dict[str, dict[str, float | int | None]]] = {}
    for fair, d in stats.items():
        out[fair] = {}
        for gk, s in d.items():
            artist_union = s.artist_img_keys | s.artist_txt_keys
            artist_match = s.artist_img_keys & s.artist_txt_keys
            exh_union = s.exh_img_keys | s.exh_txt_keys
            exh_match = s.exh_img_keys & s.exh_txt_keys
            out[fair][gk] = {
                "artist_union": len(artist_union),
                "artist_match": len(artist_match),
                "artist_rate": (100.0 * len(artist_match) / len(artist_union)) if artist_union else None,
                "artist_img_unique": len(s.artist_img_keys),
                "artist_img_count": s.artist_img_count,
                "artist_txt_unique": len(s.artist_txt_keys),
                "artist_txt_count": s.artist_txt_count,
                "exh_union": len(exh_union),
                "exh_match": len(exh_match),
                "exh_rate": (100.0 * len(exh_match) / len(exh_union)) if exh_union else None,
                "exh_img_count": s.exh_img_count,
                "exh_txt_count": s.exh_txt_count,
            }
    return out


def to_num(v: object) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def main() -> None:
    xlsx_path = Path("data/gallery_lists/rag_gellery_breakdown_master.xlsx")
    wb = load_workbook(xlsx_path)
    formal = build_formal_agg()
    run_id = datetime.now(timezone.utc).strftime("TASK_RAG_BREAKDOWN_QC_01_%Y%m%dT%H%M%SZ")
    today = datetime.now().date()

    # Fixed columns by template
    C = {
        "gallery": 1,
        "artist_union": 2,
        "artist_match": 3,
        "artist_rate": 4,
        "artist_img_unique": 5,
        "artist_img_count": 6,
        "artist_txt_unique": 7,
        "artist_txt_count": 8,
        "exh_union": 9,
        "exh_match": 10,
        "exh_rate": 11,
        "exh_img_count": 12,
        "exh_txt_count": 13,
        "memo": 14,
        "updated": 15,
        "run_id": 16,
    }

    anomaly_counts: dict[str, dict[str, int]] = {"frieze-london": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0}, "liste": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0}}
    fixed_rows: list[dict] = []
    unfixed_rows: list[dict] = []

    for sheet in ("frieze-london", "liste"):
        ws = wb[sheet]
        r = 2
        while True:
            g = ws.cell(r, C["gallery"]).value
            if g is None or str(g).strip() == "":
                break
            gallery_name = str(g)
            gk = normalize_gallery_key(gallery_name)

            art_total = to_num(ws.cell(r, C["artist_union"]).value)
            art_match = to_num(ws.cell(r, C["artist_match"]).value)
            art_img = to_num(ws.cell(r, C["artist_img_count"]).value)
            art_txt = to_num(ws.cell(r, C["artist_txt_count"]).value)
            exh_total = to_num(ws.cell(r, C["exh_union"]).value)
            exh_match = to_num(ws.cell(r, C["exh_match"]).value)
            exh_img = to_num(ws.cell(r, C["exh_img_count"]).value)
            exh_txt = to_num(ws.cell(r, C["exh_txt_count"]).value)

            A = exh_txt > 0 and exh_img == 0
            B = exh_total > 0 and exh_match == 0 and exh_img > 0 and exh_txt > 0
            Cx = art_txt > 0 and art_img == 0
            D = art_total > 0 and art_match == 0 and art_img > 0 and art_txt > 0
            E = any(ord(ch) > 127 for ch in gallery_name)
            if A:
                anomaly_counts[sheet]["A"] += 1
            if B:
                anomaly_counts[sheet]["B"] += 1
            if Cx:
                anomaly_counts[sheet]["C"] += 1
            if D:
                anomaly_counts[sheet]["D"] += 1
            if E:
                anomaly_counts[sheet]["E"] += 1

            if A or B or Cx or D:
                before = {
                    "artist_union": ws.cell(r, C["artist_union"]).value,
                    "artist_match": ws.cell(r, C["artist_match"]).value,
                    "artist_rate": ws.cell(r, C["artist_rate"]).value,
                    "artist_img_count": ws.cell(r, C["artist_img_count"]).value,
                    "artist_txt_count": ws.cell(r, C["artist_txt_count"]).value,
                    "exh_union": ws.cell(r, C["exh_union"]).value,
                    "exh_match": ws.cell(r, C["exh_match"]).value,
                    "exh_rate": ws.cell(r, C["exh_rate"]).value,
                    "exh_img_count": ws.cell(r, C["exh_img_count"]).value,
                    "exh_txt_count": ws.cell(r, C["exh_txt_count"]).value,
                }
                calc = formal.get(sheet, {}).get(gk)
                if not calc:
                    unfixed_rows.append({"sheet": sheet, "row": r, "gallery_name": gallery_name, "reason": "formal_not_found"})
                else:
                    ws.cell(r, C["artist_union"]).value = calc["artist_union"]
                    ws.cell(r, C["artist_match"]).value = calc["artist_match"]
                    ws.cell(r, C["artist_rate"]).value = calc["artist_rate"]
                    ws.cell(r, C["artist_img_unique"]).value = calc["artist_img_unique"]
                    ws.cell(r, C["artist_img_count"]).value = calc["artist_img_count"]
                    ws.cell(r, C["artist_txt_unique"]).value = calc["artist_txt_unique"]
                    ws.cell(r, C["artist_txt_count"]).value = calc["artist_txt_count"]
                    ws.cell(r, C["exh_union"]).value = calc["exh_union"]
                    ws.cell(r, C["exh_match"]).value = calc["exh_match"]
                    ws.cell(r, C["exh_rate"]).value = calc["exh_rate"]
                    ws.cell(r, C["exh_img_count"]).value = calc["exh_img_count"]
                    ws.cell(r, C["exh_txt_count"]).value = calc["exh_txt_count"]
                    ws.cell(r, C["updated"]).value = today
                    ws.cell(r, C["run_id"]).value = run_id

                    after = {
                        "artist_union": calc["artist_union"],
                        "artist_match": calc["artist_match"],
                        "artist_rate": calc["artist_rate"],
                        "artist_img_count": calc["artist_img_count"],
                        "artist_txt_count": calc["artist_txt_count"],
                        "exh_union": calc["exh_union"],
                        "exh_match": calc["exh_match"],
                        "exh_rate": calc["exh_rate"],
                        "exh_img_count": calc["exh_img_count"],
                        "exh_txt_count": calc["exh_txt_count"],
                    }
                    fixed_rows.append(
                        {
                            "sheet": sheet,
                            "row": r,
                            "gallery_name": gallery_name,
                            "triggers": [x for x, on in (("A", A), ("B", B), ("C", Cx), ("D", D), ("E", E)) if on],
                            "before": before,
                            "after": after,
                        }
                    )
            r += 1

    # Save workbook (value-only updates on anomaly rows; if none this is no-op save)
    wb.save(xlsx_path)

    md_path = Path("data/phase1_seed10/logs/rag_gallery_breakdown_qc_task_qc_01.md")
    json_path = Path("data/phase1_seed10/logs/rag_gallery_breakdown_qc_task_qc_01.json")
    md_path.parent.mkdir(parents=True, exist_ok=True)

    payload = {
        "run_id": run_id,
        "updated_at_jst": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "xlsx_path": str(xlsx_path),
        "anomaly_counts": anomaly_counts,
        "fixed_count": len(fixed_rows),
        "unfixed_count": len(unfixed_rows),
        "fixed_rows": fixed_rows,
        "unfixed_rows": unfixed_rows,
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# TASK_RAG_BREAKDOWN_QC_01",
        f"- run_id: `{run_id}`",
        f"- xlsx: `{xlsx_path}`",
        f"- fixed_count: {len(fixed_rows)}",
        f"- unfixed_count: {len(unfixed_rows)}",
        "",
        "## anomaly counts (A-D + E)",
        f"- frieze-london: {anomaly_counts['frieze-london']}",
        f"- liste: {anomaly_counts['liste']}",
        "",
        "## fixed rows (max 10)",
    ]
    for row in fixed_rows[:10]:
        lines.append(f"- {row['sheet']} row={row['row']} gallery={row['gallery_name']} triggers={row['triggers']}")
        lines.append(f"  - before: {row['before']}")
        lines.append(f"  - after:  {row['after']}")
    lines += ["", "## unfixed rows (max 10)"]
    for row in unfixed_rows[:10]:
        lines.append(f"- {row}")
    lines.append("")
    lines.append("- note: value cells only updated; styles/formatting untouched.")
    md_path.write_text("\n".join(lines), encoding="utf-8")

    print(json.dumps({"run_id": run_id, "anomaly_counts": anomaly_counts, "fixed_count": len(fixed_rows), "unfixed_count": len(unfixed_rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

