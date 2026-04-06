#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse, urlunparse
from zoneinfo import ZoneInfo

import openpyxl

from phase2_art_pulse_config import (
    get_current_artist_image_meta_paths,
    get_current_exhibitions_image_meta_paths,
    get_current_raw_paths,
)

TARGET_YEAR_DEFAULT = 2025
DEFAULT_XLSX_PATH = Path("data/gallery_lists/rag_gellery_breakdown_master.xlsx")
DEFAULT_TARGETS_CSV = Path("data/gallery_lists/phase3_fixed_block_next10_targets.csv")
SHEET_BY_FAIR = {
    "frieze_london": "frieze-london",
    "liste": "liste",
}


@dataclass
class GalleryStats:
    artist_image_keys: set[str]
    artist_image_count: int
    artist_text_keys: set[str]
    artist_text_count: int
    exhibition_image_keys: set[str]
    exhibition_image_count: int
    exhibition_text_keys: set[str]
    exhibition_text_count: int


@dataclass(frozen=True)
class ScopeTarget:
    fair_slug: str
    gallery_name_en: str

    @property
    def scope_key(self) -> tuple[str, str]:
        return (str(self.fair_slug or "").strip().casefold(), normalize_gallery_name(self.gallery_name_en))

    def to_dict(self) -> dict[str, str]:
        return {"fair_slug": self.fair_slug, "gallery_name_en": self.gallery_name_en}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Update rag_gellery_breakdown_master.xlsx from current formal artifacts for selected galleries."
    )
    parser.add_argument(
        "--targets-csv",
        default=str(DEFAULT_TARGETS_CSV),
        help=f"gallery scope CSV (default: {DEFAULT_TARGETS_CSV})",
    )
    parser.add_argument(
        "--xlsx-path",
        default=str(DEFAULT_XLSX_PATH),
        help=f"xlsx path (default: {DEFAULT_XLSX_PATH})",
    )
    parser.add_argument(
        "--target-year",
        type=int,
        default=TARGET_YEAR_DEFAULT,
        help=f"default: {TARGET_YEAR_DEFAULT}",
    )
    parser.add_argument(
        "--run-id",
        default="",
        help="run_id written to column 16; default: TASK_RAG_BREAKDOWN_01_<UTCSTAMP>",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="apply xlsx value updates; default is dry-run",
    )
    parser.add_argument(
        "--output-json",
        default="",
        help="optional JSON report path",
    )
    parser.add_argument(
        "--allow-nonstandard-scope",
        action="store_true",
        help="allow scopes other than fixed 10 (Frieze 5 + Liste 5)",
    )
    return parser.parse_args()


def utc_compact() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def normalize_gallery_name(value: str) -> str:
    text = unicodedata.normalize("NFKC", str(value or "").strip())
    text = re.sub(r"\s+", " ", text)
    text = "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")
    return text.casefold()


def normalize_url(url: str) -> str:
    value = str(url or "").strip()
    if not value:
        return ""
    try:
        parsed = urlparse(value)
    except Exception:
        return value.rstrip("/").lower()
    scheme = str(parsed.scheme or "https").lower()
    netloc = str(parsed.netloc or "").lower()
    if netloc.startswith("www."):
        netloc = netloc[4:]
    path = str(parsed.path or "").rstrip("/")
    path = path or "/"
    return urlunparse((scheme, netloc, path, "", "", ""))


def first_non_empty(*values: Any) -> str:
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return ""


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not path.exists():
        return rows
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(row, dict):
                rows.append(row)
    return rows


def fair_slug_token(value: str) -> str:
    token = str(value or "").strip().lower().replace("-", "_")
    if token in SHEET_BY_FAIR:
        return token
    return ""


def load_targets_csv(path: Path) -> dict[str, list[str]]:
    ordered_targets = load_targets_ordered(path)
    out: dict[str, list[str]] = {fair: [] for fair in SHEET_BY_FAIR}
    for target in ordered_targets:
        out[target.fair_slug].append(target.gallery_name_en)
    return out


def load_targets_ordered(path: Path) -> list[ScopeTarget]:
    if not path.exists():
        raise FileNotFoundError(f"Missing targets CSV: {path}")
    out: list[ScopeTarget] = []
    seen: set[tuple[str, str]] = set()
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            fair_slug = fair_slug_token(row.get("fair_slug") or row.get("fair") or "")
            gallery_name = str(row.get("gallery_name_en") or row.get("gallery_name") or "").strip()
            if not fair_slug or not gallery_name:
                continue
            target = ScopeTarget(fair_slug=fair_slug, gallery_name_en=gallery_name)
            if target.scope_key in seen:
                continue
            seen.add(target.scope_key)
            out.append(target)
    if not out:
        raise RuntimeError(f"No valid targets found in CSV: {path}")
    return out


def infer_artist_image_count(row: dict[str, Any]) -> int:
    for key in ("works_image_local_paths", "works_image_urls", "works_image_payload_hashes"):
        value = row.get(key)
        if isinstance(value, list):
            return len([v for v in value if str(v or "").strip()])
    try:
        return max(0, int(row.get("image_count_non_empty") or 0))
    except Exception:
        return 0


def build_stats(target_year: int) -> dict[tuple[str, str], GalleryStats]:
    stats: dict[tuple[str, str], GalleryStats] = {}

    def ensure(fair_slug: str, gallery_name: str) -> GalleryStats:
        key = (fair_slug, normalize_gallery_name(gallery_name))
        if key not in stats:
            stats[key] = GalleryStats(
                artist_image_keys=set(),
                artist_image_count=0,
                artist_text_keys=set(),
                artist_text_count=0,
                exhibition_image_keys=set(),
                exhibition_image_count=0,
                exhibition_text_keys=set(),
                exhibition_text_count=0,
            )
        return stats[key]

    for fair_slug, path in get_current_raw_paths("artists", target_year).items():
        for row in read_jsonl(path):
            gallery_name = str(row.get("gallery_name_en") or "").strip()
            if not gallery_name:
                continue
            item = ensure(fair_slug, gallery_name)
            artist_key = first_non_empty(
                row.get("artist_key"),
                row.get("artist_identity_key"),
                row.get("artist_name_key"),
                normalize_url(str(row.get("source_url") or "")),
            )
            if artist_key:
                item.artist_text_keys.add(artist_key)
            item.artist_text_count += 1

    for fair_slug, path in get_current_artist_image_meta_paths().items():
        for row in read_jsonl(path):
            gallery_name = str(row.get("gallery_name_en") or "").strip()
            if not gallery_name:
                continue
            item = ensure(fair_slug, gallery_name)
            artist_key = first_non_empty(
                row.get("artist_key"),
                row.get("artist_identity_key"),
                row.get("artist_name_key"),
                normalize_url(str(row.get("source_url") or "")),
            )
            if artist_key:
                item.artist_image_keys.add(artist_key)
            item.artist_image_count += infer_artist_image_count(row)

    for fair_slug, path in get_current_raw_paths("exhibitions", target_year).items():
        for row in read_jsonl(path):
            gallery_name = str(row.get("gallery_name_en") or "").strip()
            if not gallery_name:
                continue
            item = ensure(fair_slug, gallery_name)
            exhibition_key = first_non_empty(
                row.get("exhibition_key"),
                normalize_url(str(row.get("source_url") or "")),
            )
            if exhibition_key:
                item.exhibition_text_keys.add(exhibition_key)
            item.exhibition_text_count += 1

    for fair_slug, path in get_current_exhibitions_image_meta_paths(target_year).items():
        for row in read_jsonl(path):
            gallery_name = str(row.get("gallery_name_en") or "").strip()
            if not gallery_name:
                continue
            item = ensure(fair_slug, gallery_name)
            exhibition_key = first_non_empty(
                row.get("exhibition_key"),
                normalize_url(str(row.get("source_url") or "")),
            )
            if exhibition_key:
                item.exhibition_image_keys.add(exhibition_key)
            item.exhibition_image_count += 1

    return stats


def ratio_pct(match_count: int, total_count: int) -> float | None:
    if total_count <= 0:
        return None
    return (100.0 * float(match_count)) / float(total_count)


def sheet_existing_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[dict[str, int], int]:
    rows: dict[str, int] = {}
    first_blank = ws.max_row + 1
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        if value is None or not str(value).strip():
            first_blank = row_idx
            break
        rows[normalize_gallery_name(str(value))] = row_idx
    return rows, first_blank


def build_row_values(item: GalleryStats) -> dict[int, Any]:
    artist_union = item.artist_image_keys | item.artist_text_keys
    artist_intersection = item.artist_image_keys & item.artist_text_keys
    exhibition_union = item.exhibition_image_keys | item.exhibition_text_keys
    exhibition_intersection = item.exhibition_image_keys & item.exhibition_text_keys
    return {
        2: len(artist_union),
        3: len(artist_intersection),
        4: ratio_pct(len(artist_intersection), len(artist_union)),
        5: len(item.artist_image_keys),
        6: item.artist_image_count,
        7: len(item.artist_text_keys),
        8: item.artist_text_count,
        9: len(exhibition_union),
        10: len(exhibition_intersection),
        11: ratio_pct(len(exhibition_intersection), len(exhibition_union)),
        12: item.exhibition_image_count,
        13: item.exhibition_text_count,
    }


def read_xlsx_target_snapshot(xlsx_path: Path, targets: list[ScopeTarget]) -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    snapshots: dict[str, dict[str, Any]] = {}
    for target in targets:
        sheet_name = SHEET_BY_FAIR[target.fair_slug]
        ws = wb[sheet_name]
        row_values: dict[str, Any] = {}
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=1).value
            if normalize_gallery_name(value) != normalize_gallery_name(target.gallery_name_en):
                continue
            row_values = {str(col): ws.cell(row=row_idx, column=col).value for col in range(1, 17)}
            break
        snapshots[f"{target.fair_slug}::{target.gallery_name_en}"] = row_values
    return snapshots


def build_breakdown_update_report(
    *,
    targets: list[ScopeTarget],
    xlsx_path: Path,
    target_year: int,
    run_id: str,
    apply: bool,
    stats: dict[tuple[str, str], GalleryStats] | None = None,
) -> dict[str, Any]:
    effective_stats = stats if stats is not None else build_stats(int(target_year))
    before = read_xlsx_target_snapshot(xlsx_path, targets)
    wb = openpyxl.load_workbook(xlsx_path)
    jst_today = datetime.now(ZoneInfo("Asia/Tokyo")).date()
    updates: list[dict[str, Any]] = []
    row_action_counts: Counter[str] = Counter()
    fair_counts: Counter[str] = Counter()
    for target in targets:
        fair_counts[target.fair_slug] += 1
    for fair_slug, sheet_name in SHEET_BY_FAIR.items():
        fair_targets = [target for target in targets if target.fair_slug == fair_slug]
        if not fair_targets:
            continue
        ws = wb[sheet_name]
        existing_rows, next_row = sheet_existing_rows(ws)
        for target in fair_targets:
            norm_name = normalize_gallery_name(target.gallery_name_en)
            row_idx = existing_rows.get(norm_name)
            action = "overwrite"
            if row_idx is None:
                row_idx = next_row
                next_row += 1
                existing_rows[norm_name] = row_idx
                action = "append"
            stat = effective_stats.get((fair_slug, norm_name))
            if stat is None:
                stat = GalleryStats(set(), 0, set(), 0, set(), 0, set(), 0)
            values = build_row_values(stat)
            if apply:
                ws.cell(row=row_idx, column=1, value=target.gallery_name_en)
                for col_idx, value in values.items():
                    ws.cell(row=row_idx, column=col_idx, value=value)
                ws.cell(row=row_idx, column=15, value=jst_today)
                ws.cell(row=row_idx, column=16, value=run_id)
            updates.append(
                {
                    "fair_slug": fair_slug,
                    "sheet": sheet_name,
                    "gallery_name": target.gallery_name_en,
                    "row": row_idx,
                    "action": action,
                    "before": before.get(f"{fair_slug}::{target.gallery_name_en}", {}),
                    "after": {"1": target.gallery_name_en, **{str(k): v for k, v in values.items()}, "15": jst_today, "16": run_id},
                    "values": values,
                }
            )
            row_action_counts[action] += 1

    if apply:
        wb.save(xlsx_path)

    return {
        "mode": "apply" if apply else "dry_run",
        "xlsx_path": str(xlsx_path),
        "target_year": int(target_year),
        "run_id": run_id,
        "target_counts_by_fair": dict(fair_counts),
        "target_total": len(targets),
        "row_action_counts": dict(row_action_counts),
        "before": before,
        "updates": updates,
    }


def main() -> int:
    args = parse_args()
    targets_csv_path = Path(args.targets_csv)
    if not targets_csv_path.is_absolute():
        targets_csv_path = (Path.cwd() / targets_csv_path).resolve()
    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.is_absolute():
        xlsx_path = (Path.cwd() / xlsx_path).resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Missing xlsx: {xlsx_path}")

    ordered_targets = load_targets_ordered(targets_csv_path)
    fair_counts: Counter[str] = Counter()
    for target in ordered_targets:
        fair_counts[target.fair_slug] += 1
    total_targets = sum(fair_counts.values())
    if not args.allow_nonstandard_scope:
        if total_targets != 10 or fair_counts.get("frieze_london", 0) != 5 or fair_counts.get("liste", 0) != 5:
            raise RuntimeError(
                "fixed_scope_violation_expected_10_with_split_frieze5_liste5; "
                f"actual_total={total_targets} fair_counts={fair_counts}"
            )

    run_id = str(args.run_id or "").strip() or f"TASK_RAG_BREAKDOWN_01_{utc_compact()}"
    report = build_breakdown_update_report(
        targets=ordered_targets,
        xlsx_path=xlsx_path,
        target_year=int(args.target_year),
        run_id=run_id,
        apply=bool(args.apply),
    )
    report["targets_csv"] = str(targets_csv_path)
    report["fixed_scope_check_bypassed"] = bool(args.allow_nonstandard_scope)

    if str(args.output_json or "").strip():
        output_json_path = Path(args.output_json)
        if not output_json_path.is_absolute():
            output_json_path = (Path.cwd() / output_json_path).resolve()
        output_json_path.parent.mkdir(parents=True, exist_ok=True)
        output_json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"[breakdown-update] report={output_json_path}")
    print(
        "[breakdown-update] "
        f"mode={report['mode']} target_total={report['target_total']} "
        f"append={report['row_action_counts'].get('append', 0)} overwrite={report['row_action_counts'].get('overwrite', 0)}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
